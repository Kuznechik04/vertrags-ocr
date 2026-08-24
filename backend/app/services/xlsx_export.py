"""Baut Excel-Arbeitsmappen (.xlsx) aus den erkannten/validierten Vertragsfeldern.

Genutzt vom Export-Button im Frontend – sowohl für ein einzelnes Dokument als
auch für einen gesammelten Export mehrerer Dokumente (z.B. alle eigenen
Verträge bzw. für Admins alle Verträge aller Nutzer).
"""
from __future__ import annotations

import json
from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

if TYPE_CHECKING:
    # Nur für Typannotationen nötig – vermeidet, dass dieses Modul zur Laufzeit
    # zwingend SQLAlchemy-Modelle importieren muss.
    from app.models.document import Document

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = "2F6FED"


def _collect_field_columns(documents: list[Document]) -> list[tuple[str, str]]:
    """Sammelt die (field_key, field_label)-Paare, die in den zu exportierenden
    Dokumenten tatsächlich vorkommen, in Reihenfolge des ersten Auftretens.
    Läuft dokumentübergreifend, da unterschiedliche Vertragstyp-Templates
    unterschiedliche Feldsets haben können – die Kopfzeile deckt die Union ab,
    pro Dokument bleiben nicht zutreffende Spalten leer."""
    columns: list[tuple[str, str]] = []
    seen: set[str] = set()
    for document in documents:
        for field in document.fields:
            if field.field_key not in seen:
                seen.add(field.field_key)
                columns.append((field.field_key, field.field_label))
    return columns


def _style_header(ws: Worksheet, num_columns: int) -> None:
    from openpyxl.styles import PatternFill

    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _autosize_columns(ws: Worksheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _field_values_and_details(document: Document, field_keys: list[str]) -> tuple[list[str], dict]:
    """Liefert pro Dokument die finalen Feldwerte (in `field_keys`-Reihenfolge,
    für die Feld-Spalten) sowie ein Detail-Dict mit Konfidenz/Validierungs-
    Checks pro Feld (wird als JSON in einer einzigen Spalte ausgegeben)."""
    fields_by_key = {f.field_key: f for f in document.fields}
    values: list[str] = []
    details: dict = {}
    for key in field_keys:
        field = fields_by_key.get(key)
        values.append((field.final_value or "") if field else "")
        if field:
            details[key] = {
                "konfidenz": round(field.confidence, 2),
                "validiert": field.is_validated,
                "korrigiert": bool(field.is_corrected or field.is_position_corrected),
            }
    return values, details


def build_single_document_xlsx(document: Document) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vertragsfelder"

    columns = _collect_field_columns([document])
    field_keys = [key for key, _ in columns]
    field_labels = [label for _, label in columns]

    headers = field_labels + ["Details (JSON)"]
    ws.append(headers)
    _style_header(ws, len(headers))

    values, details = _field_values_and_details(document, field_keys)
    ws.append(values + [json.dumps(details, ensure_ascii=False)])

    _autosize_columns(ws, [24] * len(field_labels) + [60])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_multi_document_xlsx(documents: list[Document]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vertragsfelder"

    columns = _collect_field_columns(documents)
    field_keys = [key for key, _ in columns]
    field_labels = [label for _, label in columns]

    headers = ["Dateiname", "Hochgeladen von", "Status"] + field_labels + ["Details (JSON)"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for document in documents:
        values, details = _field_values_and_details(document, field_keys)
        ws.append(
            [
                document.filename,
                document.owner_email or document.owner_id,
                document.status.value if hasattr(document.status, "value") else document.status,
            ]
            + values
            + [json.dumps(details, ensure_ascii=False)]
        )

    _autosize_columns(ws, [28, 24, 14] + [24] * len(field_labels) + [60])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
